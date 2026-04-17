"""
Excel exporter for rate card and proposal history.
Produces a .xlsx with 3 worksheets: Platform Stats, Rate Card, Proposal History.
"""
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from ratecard.core.models import (
    AFFILIATE_PLATFORM_LABELS,
    CONTENT_TYPE_LABELS,
    CONTENT_TYPE_PLATFORM,
    PLATFORM_LABELS,
    CreatorProfile,
    Package,
    Proposal,
)
from ratecard.core.pricing import calculate_affiliate_rate, format_idr


def _hex_fill(hex_color: str) -> PatternFill:
    color = hex_color.lstrip("#")
    return PatternFill("solid", fgColor=color)


def _thin_border() -> Border:
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_style(ws, row: int, cols: int, fill_hex: str, font_color: str = "FFFFFF"):
    fill = _hex_fill(fill_hex)
    font = Font(bold=True, color=font_color)
    border = _thin_border()
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)


def export_xlsx(
    profile: CreatorProfile,
    packages: List[Package],
    proposals: List[Proposal],
    output_path: Path,
) -> Path:
    """Export rate card data to Excel. Returns the output path."""
    wb = Workbook()
    brand = (profile.brand_color_hex or "#E91E8C").lstrip("#")

    # ── Sheet 1: Platform Stats ──────────────────────────────
    ws1 = wb.active
    ws1.title = "Platform Stats"

    headers = ["Platform", "Followers", "Avg Views", "Engagement Rate (%)",
               "Avg Monthly Impressions", "Last Updated"]
    ws1.append(headers)
    _header_style(ws1, 1, len(headers), brand)

    alt_fill = _hex_fill("FFF5FB")
    for i, s in enumerate(profile.platform_stats):
        if s.followers == 0:
            continue
        row = [
            PLATFORM_LABELS.get(s.platform, s.platform.value),
            s.followers,
            s.avg_views,
            round(s.engagement_rate * 100, 2),
            s.avg_monthly_impressions,
            s.last_updated.strftime("%d/%m/%Y") if s.last_updated else "",
        ]
        ws1.append(row)
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws1.cell(row=i + 2, column=col).fill = alt_fill
        for col in range(1, len(headers) + 1):
            ws1.cell(row=i + 2, column=col).border = _thin_border()

    ws1.freeze_panes = "A2"
    _autofit(ws1)

    # ── Sheet 2: Rate Card ───────────────────────────────────
    ws2 = wb.create_sheet("Rate Card")

    headers2 = ["Tier", "Konten", "Platform", "Qty", "Usage Rights (hari)",
                "Exclusivity (hari)", "Harga Satuan (IDR)", "Subtotal (IDR)",
                "Bundle Discount (%)", "Final Package Price (IDR)", "Valid (hari)"]
    ws2.append(headers2)
    _header_style(ws2, 1, len(headers2), brand)

    row_num = 2
    for pkg in packages:
        for j, item in enumerate(pkg.items):
            platform = CONTENT_TYPE_PLATFORM.get(item.content_type)
            row = [
                pkg.tier.value,
                CONTENT_TYPE_LABELS.get(item.content_type, item.content_type.value),
                PLATFORM_LABELS.get(platform, "") if platform else "",
                item.quantity,
                item.usage_rights_days if item.usage_rights_days else "",
                item.exclusivity_days if item.exclusivity_days else "",
                int(item.unit_price),
                int(item.unit_price * item.quantity),
                f"{pkg.bundle_discount_pct * 100:.0f}%" if j == 0 else "",
                int(pkg.final_price) if j == 0 else "",
                pkg.valid_days if j == 0 else "",
            ]
            ws2.append(row)
            if row_num % 2 == 0:
                for col in range(1, len(headers2) + 1):
                    ws2.cell(row=row_num, column=col).fill = alt_fill
            for col in range(1, len(headers2) + 1):
                ws2.cell(row=row_num, column=col).border = _thin_border()

            # Bold final price column
            if j == 0:
                ws2.cell(row=row_num, column=10).font = Font(bold=True, color=brand)

            row_num += 1

        # Empty row between tiers
        ws2.append([])
        row_num += 1

    ws2.freeze_panes = "A2"
    _autofit(ws2)

    # ── Sheet 3: Proposal History ────────────────────────────
    ws3 = wb.create_sheet("Proposal History")

    headers3 = ["ID", "Klien", "Perusahaan", "Email", "Campaign",
                "Pricing Model", "Paket", "Total (IDR)", "Discount (%)", "Status",
                "Dibuat", "Dikirim", "PDF Path", "Catatan"]
    ws3.append(headers3)
    _header_style(ws3, 1, len(headers3), brand)

    status_colors = {
        "draft":    "CCCCCC",
        "sent":     "AED6F1",
        "accepted": "A9DFBF",
        "rejected": "F1948A",
    }

    for i, p in enumerate(proposals):
        pkg_names = ", ".join(pkg.tier.value for pkg in p.packages)
        pm_label = p.pricing_model.value.title() if hasattr(p, 'pricing_model') else "Flat"
        row = [
            p.id,
            p.client_name,
            p.client_company,
            p.client_email,
            p.campaign_name,
            pm_label,
            pkg_names,
            int(p.total_price),
            f"{p.discount_pct * 100:.0f}%" if p.discount_pct else "0%",
            p.status.value.upper(),
            p.created_at.strftime("%d/%m/%Y") if p.created_at else "",
            p.sent_at.strftime("%d/%m/%Y") if p.sent_at else "",
            p.pdf_path or "",
            p.notes,
        ]
        ws3.append(row)
        row_num3 = i + 2

        # Color status cell (column 10 after adding pricing_model)
        status_hex = status_colors.get(p.status.value, "FFFFFF")
        ws3.cell(row=row_num3, column=10).fill = _hex_fill(status_hex)
        ws3.cell(row=row_num3, column=10).font = Font(bold=True)

        # Color total cell (column 8 after adding pricing_model)
        ws3.cell(row=row_num3, column=8).font = Font(bold=True, color=brand)

        for col in range(1, len(headers3) + 1):
            ws3.cell(row=row_num3, column=col).border = _thin_border()
            if i % 2 == 0:
                if col != 10:  # don't override status color
                    ws3.cell(row=row_num3, column=col).fill = alt_fill

    ws3.freeze_panes = "A2"
    _autofit(ws3)

    # ── Sheet 4: Affiliate Stats + Commission Rates ─────────
    active_aff = [a for a in profile.affiliate_stats if a.enabled and a.avg_monthly_gmv > 0]
    if active_aff:
        ws_aff = wb.create_sheet("Affiliate")

        # Stats section
        aff_headers = ["Platform", "Avg Monthly GMV (IDR)", "Conversion Rate (%)",
                       "Avg Commission (%)", "Last Updated"]
        ws_aff.append(aff_headers)
        _header_style(ws_aff, 1, len(aff_headers), brand)

        for i, aff in enumerate(active_aff):
            row = [
                AFFILIATE_PLATFORM_LABELS.get(aff.platform, aff.platform.value),
                int(aff.avg_monthly_gmv),
                round(aff.avg_conversion_rate * 100, 2),
                round(aff.avg_commission_pct * 100, 2),
                aff.last_updated.strftime("%d/%m/%Y") if aff.last_updated else "",
            ]
            ws_aff.append(row)
            row_idx = i + 2
            for col in range(1, len(aff_headers) + 1):
                ws_aff.cell(row=row_idx, column=col).border = _thin_border()
                if i % 2 == 0:
                    ws_aff.cell(row=row_idx, column=col).fill = alt_fill
            ws_aff.cell(row=row_idx, column=2).font = Font(bold=True, color=brand)

        # Blank row separator
        ws_aff.append([])
        sug_start = len(active_aff) + 3

        # Commission suggestions section
        sug_headers = ["Platform", "Komisi Min (%)", "Komisi Target (%)", "Komisi Max (%)",
                       "Projected Earning Min (IDR)", "Projected Earning Target (IDR)",
                       "Base Fee Suggested (IDR)", "Min Campaign GMV (IDR)"]
        ws_aff.append(sug_headers)
        _header_style(ws_aff, sug_start, len(sug_headers), brand)

        for i, aff in enumerate(active_aff):
            sug = calculate_affiliate_rate(
                aff.platform.value, aff.avg_monthly_gmv,
                aff.avg_conversion_rate, aff.avg_commission_pct,
            )
            row = [
                AFFILIATE_PLATFORM_LABELS.get(aff.platform, aff.platform.value),
                round(sug["commission_min_pct"] * 100, 1),
                round(sug["commission_target_pct"] * 100, 1),
                round(sug["commission_max_pct"] * 100, 1),
                int(sug["projected_monthly_earning_min"]),
                int(sug["projected_monthly_earning_target"]),
                int(sug["base_fee_suggested"]),
                int(sug["min_campaign_value"]),
            ]
            ws_aff.append(row)
            row_idx = sug_start + i + 1
            for col in range(1, len(sug_headers) + 1):
                ws_aff.cell(row=row_idx, column=col).border = _thin_border()
                if i % 2 == 0:
                    ws_aff.cell(row=row_idx, column=col).fill = alt_fill
            # Bold target commission
            ws_aff.cell(row=row_idx, column=3).font = Font(bold=True, color=brand)
            ws_aff.cell(row=row_idx, column=6).font = Font(bold=True, color=brand)

        if profile.affiliate_categories:
            ws_aff.append([])
            ws_aff.append(["Kategori Unggulan", " · ".join(profile.affiliate_categories)])

        ws_aff.freeze_panes = "A2"
        _autofit(ws_aff)

    # ── Meta sheet ───────────────────────────────────────────
    ws4 = wb.create_sheet("Info")
    ws4.append(["Rate Card Manager"])
    ws4.append(["Creator", profile.name])
    ws4.append(["Niche", profile.niche])
    ws4.append(["Generated", datetime.now().strftime("%d %B %Y %H:%M")])
    ws4["A1"].font = Font(bold=True, size=14, color=brand)

    wb.save(str(output_path))
    return output_path
